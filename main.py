"""
Packing Template Generator
-------------------------
A Tkinter-based desktop application for generating ITP and Packing List templates.
Uses `openpyxl` to copy an XLSX template and fill in user-supplied data.

This is an MVP implementation that demonstrates the full flow described in
`brief.md` and `flowchart.md`.  It focuses on Packing List templates but is
structured to allow easy addition of further sub-templates or template types.

Author: Cascade AI assistant
"""

from __future__ import annotations

import datetime as _dt
import shutil as _shutil
import sys as _sys
from pathlib import Path
from tkinter import (
    BOTH,
    END,
    LEFT,
    RIGHT,
    StringVar,
    Tk,
    filedialog,
    messagebox,
)
from tkinter import ttk  # themed widgets

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover – helpful error if dependency missing
    print("The 'openpyxl' package is required. Install with: pip install openpyxl")
    _sys.exit(1)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

PROJECT_DIR = Path(__file__).resolve().parent

TEMPLATES: dict[str, dict[str, Path]] = {
    "ITP": {
        "33kV Disc Manual + Single E/S Manual": PROJECT_DIR / "ITP LIST -Template.xls",  # noqa: E501 – placeholder
        "72.5kV Disc Motorised + Dual E/S Motorised": PROJECT_DIR / "ITP LIST -Template.xls",  # noqa: E501
    },
    "Packing List": {
        "72.5kV Disc Motorised + Dual E/S Motorised": PROJECT_DIR
        / "PACKING LIST - Disc 72.5kV Motor Dual ES Motor.xlsx",
        "33kV Disc Manual + Single E/S Manual": PROJECT_DIR
        / "PACKING LIST - Disc 33kV Manual Single ES Manual.xlsx",
    },
}

DATE_FMT = "%d/%m/%Y"  # Excel friendly date format

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _validate_entries(values: dict[str, str]) -> bool:
    """Return True if all required fields are non-empty, else show an error."""
    missing = [k for k, v in values.items() if not v.strip()]
    if missing:
        messagebox.showerror("Missing Data", f"Please fill in: {', '.join(missing)}")
        return False
    return True


def _fill_workbook(template_path: Path, dest_path: Path, values: dict[str, str]) -> None:
    """Copy `template_path` to `dest_path` and write values to known cells.

    NOTE: Adjust the cell mappings below to match real templates.
    """
    # First copy the file to preserve formatting and extra sheets
    _shutil.copy(template_path, dest_path)

    wb = load_workbook(dest_path)
    ws = wb.active  # assume primary sheet

    def _write_safe(coord: str, val: str) -> None:
        """Write *val* to *coord* handling merged cells gracefully."""
        try:
            ws[coord].value = val
        except AttributeError:
            # coord might refer to a read-only MergedCell; find its master cell
            for rng in ws.merged_cells.ranges:
                if coord in rng:
                    master_coord = rng.start_cell.coordinate  # top-left
                    ws[master_coord].value = val
                    break
            else:
                raise  # re-raise if not part of any merged range

    # Mapping: {column_title: cell_reference}
    cell_map = {
        # Write only to column C, rows 4–8 as requested by the user
        "Customer": "C4",
        "Purchase Order": "C5",
        "Date": "C6",
        "Drawing No.": "C7",
        "Serial No.": "C8",
    }

    for field, cell in cell_map.items():
        if field in values:
            _write_safe(cell, values[field])

    wb.save(dest_path)


# ---------------------------------------------------------------------------
# GUI Application
# ---------------------------------------------------------------------------

class PackingTemplateApp(Tk):
    """Main application window."""

    def __init__(self) -> None:
        super().__init__()

        self.title("Packing Template Generator")
        self.geometry("500x400")

        self._create_variables()
        self._build_gui()

    # ------------------ GUI construction helpers ------------------ #

    def _create_variables(self) -> None:
        self.template_type = StringVar(value="Packing List")
        self.equipment = StringVar()
        self.customer = StringVar()
        self.purchase_order = StringVar()
        self.date = StringVar(value=_dt.datetime.today().strftime(DATE_FMT))
        self.drawing_no = StringVar()
        self.serial_no = StringVar()

    def _build_gui(self) -> None:
        pad_opts = {"padx": 10, "pady": 5}

        # Template type selector
        ttk.Label(self, text="Template Type:").pack(anchor="w", **pad_opts)
        cb_template = ttk.Combobox(
            self,
            textvariable=self.template_type,
            state="readonly",
            values=list(TEMPLATES.keys()),
        )
        cb_template.pack(fill="x", **pad_opts)
        cb_template.bind("<<ComboboxSelected>>", self._on_template_change)

        # Equipment selector
        ttk.Label(self, text="Equipment:").pack(anchor="w", **pad_opts)
        self.cb_equipment = ttk.Combobox(self, textvariable=self.equipment, state="readonly")
        self.cb_equipment.pack(fill="x", **pad_opts)

        # Initial equipment list
        self._refresh_equipment_options()

        # Text fields in a simple grid
        form_frame = ttk.Frame(self)
        form_frame.pack(fill=BOTH, expand=True, **pad_opts)

        def _add_field(label_text: str, var: StringVar, row: int) -> None:
            ttk.Label(form_frame, text=f"{label_text}:").grid(row=row, column=0, sticky="w", padx=5, pady=5)
            entry = ttk.Entry(form_frame, textvariable=var)
            entry.grid(row=row, column=1, sticky="ew", padx=5, pady=5)
            form_frame.columnconfigure(1, weight=1)

        _add_field("Customer", self.customer, 0)
        _add_field("Purchase Order", self.purchase_order, 1)
        _add_field("Date", self.date, 2)
        _add_field("Drawing No.", self.drawing_no, 3)
        _add_field("Serial No.", self.serial_no, 4)

        # Export button
        ttk.Button(self, text="Export", command=self._export).pack(side=RIGHT, **pad_opts)

    # ------------------ Callbacks ------------------ #

    def _on_template_change(self, _event=None) -> None:  # noqa: D401 – Tkinter callback
        """Refresh equipment options when template type changes."""
        self._refresh_equipment_options()

    def _refresh_equipment_options(self) -> None:
        tmpl = self.template_type.get()
        options = list(TEMPLATES[tmpl].keys())
        self.cb_equipment["values"] = options
        # Keep previously selected equipment if still valid, else reset
        if self.equipment.get() not in options:
            self.equipment.set(options[0])

    def _gather_form_data(self) -> dict[str, str]:
        return {
            "Template Type": self.template_type.get(),
            "Equipment": self.equipment.get(),
            "Customer": self.customer.get(),
            "Purchase Order": self.purchase_order.get(),
            "Date": self.date.get() or _dt.datetime.today().strftime(DATE_FMT),
            "Drawing No.": self.drawing_no.get(),
            "Serial No.": self.serial_no.get(),
        }

    # ------------------ Export logic ------------------ #

    def _export(self) -> None:
        data = self._gather_form_data()
        if not _validate_entries({k: v for k, v in data.items() if k not in {"Template Type"}}):
            return

        # Ask for save location
        dest_path_str = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")],
            initialfile=f"{data['Template Type'].replace(' ', '_')}_{data['Equipment'].split()[0]}_{_dt.datetime.now().strftime('%Y%m%d')}.xlsx",
        )
        if not dest_path_str:
            return  # User cancelled

        dest_path = Path(dest_path_str)

        # Locate template
        template_path = TEMPLATES[data["Template Type"]][data["Equipment"]]
        if not template_path.exists():
            messagebox.showerror("Template Missing", f"Template file not found:\n{template_path}")
            return

        try:
            _fill_workbook(template_path, dest_path, data)
        except Exception as exc:  # pragma: no cover – generic catch for user feedback
            messagebox.showerror("Export Error", f"Failed to create file:\n{exc}")
            return

        messagebox.showinfo("Success", f"File saved to:\n{dest_path}")


# ---------------------------------------------------------------------------
# Entry-point
# ---------------------------------------------------------------------------

def main() -> None:  # noqa: D401 – conventional main
    app = PackingTemplateApp()
    app.mainloop()


if __name__ == "__main__":
    main()
